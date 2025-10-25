from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Any, Optional

import pandas as pd
import warnings

from order_shipping_status.models import EnvCfg
from order_shipping_status.pipelines.column_contract import ColumnContract
from order_shipping_status.pipelines.enricher import Enricher
from order_shipping_status.pipelines.preprocessor import Preprocessor
from order_shipping_status.rules.indicators import apply_indicators
from order_shipping_status.rules.status_mapper import map_indicators_to_status
from openpyxl import load_workbook


class WorkbookProcessor:
    """Orchestrates pre-processing, column contract, enrichment, and rules."""

    def __init__(
        self,
        logger,
        *,
        client: Optional[Any] = None,
        normalizer: Optional[Any] = None,
        reference_date: dt.date | None = None,
        enable_date_filter: bool = True,
        stalled_threshold_days: int = 4,
        reference_now: dt.datetime | None = None,
    ) -> None:
        self.logger = logger
        self.client = client
        self.normalizer = normalizer
        self.reference_date = reference_date
        self.enable_date_filter = enable_date_filter
        self.stalled_threshold_days = int(stalled_threshold_days)
        self.reference_now = reference_now

    def process(
        self,
        input_path: Path,
        processed_path: Path,
        env_cfg: Optional[EnvCfg] = None,
        *,
        sidecar_dir: Optional[Path] = None,
    ) -> dict[str, Any]:
        input_path = Path(input_path)
        processed_path = Path(processed_path)

        if not input_path.exists():
            self.logger.error("Input file does not exist: %s", input_path)
            raise FileNotFoundError(input_path)

        df_in = self._read_input(input_path)

        df_out = self._prepare_and_enrich(df_in, sidecar_dir=sidecar_dir)
        # Optional: developer preview of a few columns (only those that exist)
        try:
            _want = [
                "Tracking Number",
                "derivedCode",
                "IsPreTransit",
                "IsDelivered",
                "ScanEventsCount",
                "LatestEventTimestampUtc",
                "DaysSinceLatestEvent",
                "IsStalled",
            ]
            _have = [c for c in _want if c in df_out.columns]
            if _have:
                print(df_out[_have].to_string(index=False))
        except Exception:
            pass

        now_utc = dt.datetime.now(dt.timezone.utc).isoformat()
        has_creds = bool(
            getattr(env_cfg, "SHIPPING_CLIENT_ID", "")
            and getattr(env_cfg, "SHIPPING_CLIENT_SECRET", "")
        )

        marker = self._build_marker(
            input_path, processed_path, now_utc, has_creds, df_in, df_out)

        # Ensure output dir exists
        processed_path.parent.mkdir(parents=True, exist_ok=True)

        # Write workbook (All Shipments, All Issues, PreTransit, Stalled, Processed, Marker)
        self._write_workbook(processed_path, df_in, df_out, marker)

        # Post-process workbook (fix empty-string cells in 'Processed')
        self._postprocess_workbook(processed_path)

        self.logger.info("Wrote processed workbook → %s", processed_path)
        return {
            "output_path": str(processed_path),
            "env_has_creds": has_creds,
            "timestamp_utc": now_utc,
            "output_cols": list(df_out.columns),
            "output_shape": (len(df_out), len(df_out.columns)),
        }

    def _read_input(self, input_path: Path) -> pd.DataFrame:
        try:
            df_in = pd.read_excel(input_path, sheet_name=0, engine="openpyxl")
            self.logger.debug(
                "Opened input workbook: %s (rows=%d, cols=%d)",
                input_path.name,
                len(df_in),
                len(df_in.columns),
            )
        except Exception as ex:
            self.logger.warning(
                "Could not read input workbook (%s): %s", input_path.name, ex
            )
            df_in = pd.DataFrame()
        return df_in

    def _prepare_and_enrich(self, df_in: pd.DataFrame, *, sidecar_dir: Optional[Path] = None) -> pd.DataFrame:
        # Preprocess  contract  enrichment
        df_prep = Preprocessor(
            self.reference_date,
            logger=self.logger,
            enable_date_filter=self.enable_date_filter,
        ).prepare(df_in)

        df_out = ColumnContract().ensure(df_prep)

        df_out = Enricher(
            self.logger,
            client=self.client,
            normalizer=self.normalizer,
        ).enrich(df_out, sidecar_dir=sidecar_dir)

        # --- Metrics: DaysSinceLatestEvent (vectorized, NaT-safe) ---
        now = pd.Timestamp(
            self.reference_now or dt.datetime.now(dt.timezone.utc))

        if "LatestEventTimestampUtc" in df_out.columns:
            ts = pd.to_datetime(
                df_out["LatestEventTimestampUtc"].astype("string"),
                errors="coerce",
                utc=True,
            )
            days = (now - ts).dt.days  # float with NaN for NaT
            df_out["DaysSinceLatestEvent"] = (
                pd.to_numeric(days, errors="coerce").fillna(0).astype("int64")
            )
        else:
            df_out["DaysSinceLatestEvent"] = 0

        # Optional debug peek (safe): only log if the columns exist
        try:
            if "latestStatusDetail" in df_out.columns and len(df_out) > 0:
                sample = df_out["latestStatusDetail"].iloc[0]
                if self.logger:
                    self.logger.debug("sample.latestStatusDetail keys: %s",
                                      list(sample.keys()) if isinstance(sample, dict) else type(sample))
            elif "raw" in df_out.columns and len(df_out) > 0:
                sample_raw = df_out["raw"].iloc[0]
                if self.logger:
                    if isinstance(sample_raw, dict):
                        self.logger.debug(
                            "sample.raw top-level keys: %s", list(sample_raw.keys()))
                    else:
                        self.logger.debug(
                            "sample.raw type: %s", type(sample_raw))
        except Exception:
            # swallow any debug inspection errors
            pass

        # Apply indicators and map to status
        df_out = apply_indicators(
            df_out, stalled_threshold_days=self.stalled_threshold_days)
        df_out = map_indicators_to_status(df_out)

        # Normalize CalculatedReasons to concrete empty strings (object dtype)
        if "CalculatedReasons" in df_out.columns:
            cr = df_out["CalculatedReasons"].astype("object")
            cr = cr.where(pd.notna(cr), "")
            df_out["CalculatedReasons"] = cr

        return df_out

    def _build_marker(self, input_path: Path, processed_path: Path, now_utc: str, has_creds: bool, df_in: pd.DataFrame, df_out: pd.DataFrame) -> pd.DataFrame:
        api_bodies = None
        try:
            # Prefer writer.path exposed by LiveFedExAdapter (writer.path is a Path)
            if hasattr(self.client, "_writer") and getattr(self.client, "_writer") is not None:
                try:
                    api_bodies = str(self.client._writer.path)
                except Exception:
                    api_bodies = None
            # Back-compat: older FedEx client used _save_bodies_path
            elif hasattr(self.client, "_save_bodies_path") and self.client._save_bodies_path:
                api_bodies = str(self.client._save_bodies_path)
        except Exception:
            api_bodies = None

        return pd.DataFrame(
            [
                {
                    "_oss_marker": "ok",
                    "input_name": input_path.name,
                    "input_dir": str(input_path.parent),
                    "output_name": processed_path.name,
                    "timestamp_utc": now_utc,
                    "env_has_creds": has_creds,
                    "api_bodies_path": api_bodies,
                    "input_rows": len(df_in),
                    "input_cols": len(df_in.columns),
                    "output_rows": len(df_out),
                    "output_cols": len(df_out.columns),
                }
            ]
        )

    def _write_workbook(self, processed_path: Path, df_in: pd.DataFrame, df_out: pd.DataFrame, marker: pd.DataFrame) -> None:
        # ---- helpers ------------------------------------------------------------
        def _clean_tn_value(v) -> str:
            """Return a clean string tracking number (no decimals/scientific)."""
            if v is None:
                return ""
            s = str(v).strip()
            if s == "" or s.lower() in ("nan", "none"):
                return ""
            # common cases: 1.2345E+11, 123456789012.0, ints/floats
            try:
                # if it's numeric in any form, render as integer with no decimal
                as_float = float(s.replace(",", ""))  # allow accidental commas
                # guard against scientific strings like '1.234e+11' -> int ok
                return str(int(as_float))
            except Exception:
                # not numeric -> keep as-is
                # but if it's like '123456789012.0', strip trailing '.0'
                if s.endswith(".0"):
                    return s[:-2]
                return s

        def _format_tracking_number_col(df: pd.DataFrame) -> pd.DataFrame:
            if "Tracking Number" not in df.columns:
                return df
            out = df.copy()
            out["Tracking Number"] = (
                out["Tracking Number"]
                .astype("object")  # don't let pandas coerce back to numeric
                .map(_clean_tn_value)
            )
            return out

        def _ensure_column_after(df: pd.DataFrame, col: str, after_col: str, default_value=0) -> pd.DataFrame:
            out = df.copy()
            if col not in out.columns:
                out[col] = default_value
            if after_col in out.columns:
                cols = list(out.columns)
                if col in cols:
                    cols.remove(col)
                insert_at = cols.index(after_col) + 1
                cols.insert(insert_at, col)
                out = out[cols]
            return out

        def _finalize(df: pd.DataFrame) -> pd.DataFrame:
            out = _format_tracking_number_col(df)
            if "Damaged" in out.columns:
                out = _ensure_column_after(
                    out, "UnableToDeliver", "Damaged", default_value=0)
            return out

        # ---- derive views -------------------------------------------------------
        if "IsPreTransit" in df_out.columns:
            pretransit = df_out[df_out["IsPreTransit"] == True]
        else:
            pretransit = pd.DataFrame(columns=df_out.columns)

        if "IsStalled" in df_out.columns:
            stalled = df_out[df_out["IsStalled"] == 1]
        else:
            stalled = pd.DataFrame(columns=df_out.columns)

        if "Damaged" in df_out.columns or "IsRTS" in df_out.columns:
            damaged_or_returned = df_out[
                ((df_out["Damaged"] == 1)
                 if "Damaged" in df_out.columns else False)
                | ((df_out["IsRTS"] == 1) if "IsRTS" in df_out.columns else False)
            ]
        else:
            damaged_or_returned = pd.DataFrame(columns=df_out.columns)

        if "IsDelivered" in df_out.columns:
            try:
                all_issues = df_out[df_out["IsDelivered"] != 1]
            except Exception:
                all_issues = pd.DataFrame(columns=df_out.columns)
        else:
            has_exc = df_out["HasException"] == 1 if "HasException" in df_out.columns else pd.Series([
                                                                                                     False] * len(df_out))
            is_stalled = df_out["IsStalled"] == 1 if "IsStalled" in df_out.columns else pd.Series([
                                                                                                  False] * len(df_out))
            is_rts = df_out["IsRTS"] == 1 if "IsRTS" in df_out.columns else pd.Series([
                                                                                      False] * len(df_out))
            try:
                issues_mask = (has_exc | is_stalled | is_rts)
                issues_mask.index = df_out.index
                all_issues = df_out[issues_mask]
            except Exception:
                all_issues = pd.DataFrame(columns=df_out.columns)

        # finalize frames
        df_in_w = _finalize(df_in)
        all_issues_w = _finalize(all_issues)
        pretransit_w = _finalize(pretransit)
        stalled_w = _finalize(stalled)
        damaged_or_returned_w = _finalize(damaged_or_returned)

        # ---- write --------------------------------------------------------------
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            with pd.ExcelWriter(processed_path, engine="openpyxl", mode="w") as xw:
                try:
                    df_in_w.to_excel(
                        xw, sheet_name="All Shipments", index=False, na_rep="")
                except Exception:
                    pd.DataFrame().to_excel(xw, sheet_name="All Shipments", index=False, na_rep="")

                all_issues_w.to_excel(
                    xw, sheet_name="All Issues", index=False, na_rep="")
                pretransit_w.to_excel(
                    xw, sheet_name="PreTransit", index=False, na_rep="")
                stalled_w.to_excel(xw, sheet_name="Stalled",
                                   index=False, na_rep="")
                damaged_or_returned_w.to_excel(
                    xw, sheet_name="Damaged or Returned", index=False, na_rep="")
                marker.to_excel(xw, sheet_name="Marker", index=False)

        # ---- force Excel TEXT type for "Tracking Number" ------------------------
        try:
            wb = load_workbook(processed_path)
            for sheet_name in ["All Shipments", "All Issues", "PreTransit", "Stalled", "Damaged or Returned"]:
                if sheet_name not in wb.sheetnames:
                    continue
                ws = wb[sheet_name]
                if ws.max_row < 1:
                    continue
                # find TN column by header
                tn_col_idx = None
                for col_idx, cell in enumerate(ws[1], start=1):
                    if (cell.value or "").strip() == "Tracking Number":
                        tn_col_idx = col_idx
                        break
                if tn_col_idx is None:
                    continue
                # coerce every TN cell to string and set number format to text
                for r in range(2, ws.max_row + 1):
                    c = ws.cell(row=r, column=tn_col_idx)
                    c.value = _clean_tn_value(c.value)
                    c.number_format = "@"  # Excel 'Text' format
            wb.save(processed_path)
        except Exception:
            # non-fatal if anything goes sideways here
            pass

    def _postprocess_workbook(self, processed_path: Path) -> None:
        try:
            wb = load_workbook(processed_path)
            # Normalize CalculatedReasons to explicit empty-string cells on any
            # sheet that contains that header (we no longer write a 'Processed'
            # sheet by default).
            for name in wb.sheetnames:
                ws = wb[name]
                header = [c.value for c in next(
                    ws.iter_rows(min_row=1, max_row=1))]
                if "CalculatedReasons" in header:
                    col_idx = header.index("CalculatedReasons")
                    for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                        cell = row[0]
                        if cell.value is None:
                            cell.value = ""
                            cell.data_type = "s"
            # Suppress any openpyxl warnings during save (non-fatal cosmetic warnings)
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                wb.save(processed_path)
        except Exception:
            # If openpyxl post-processing fails, continue — writing already completed.
            pass
