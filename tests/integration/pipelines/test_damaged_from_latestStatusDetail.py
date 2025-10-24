from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
import pytest
import logging

from order_shipping_status.pipelines.workbook_processor import WorkbookProcessor

SAMPLE_BODY = {
    "transactionId": "79851028-9cf1-422e-b276-8164ba017f29",
    "output": {
        "completeTrackResults": [
            {
                "trackingNumber": "394273834920",
                "trackResults": [
                    {
                        "trackingNumberInfo": {
                            "trackingNumber": "394273834920",
                            "trackingNumberUniqueId": "12029~394273834920~FDEG",
                            "carrierCode": "FDXG"
                        },
                        "latestStatusDetail": {
                            "code": "DE",
                            "derivedCode": "DE",
                            "statusByLocale": "Delivery exception",
                            "description": "Delivery exception",
                            "scanLocation": {
                                "city": "COCOA",
                                "stateOrProvinceCode": "FL",
                                "countryCode": "US",
                                "residential": False,
                                "countryName": "United States"
                            },
                            "ancillaryDetails": [
                                {
                                    "reason": "039",
                                    "reasonDescription": "Damaged, handling per shipper instructions",
                                    "action": "Please contact your shipper for more information.",
                                    "actionDescription": "Damaged - handling per shipper instructions"
                                }
                            ]
                        },
                        "scanEvents": []
                    }
                ]
            }
        ]
    }
}


def _normalized_row_from_sample() -> dict:
    tr = SAMPLE_BODY["output"]["completeTrackResults"][0]["trackResults"][0]
    lsd = tr["latestStatusDetail"]
    return {
        # columns your pipeline relies on
        "Tracking Number": tr["trackingNumberInfo"]["trackingNumber"],
        "derivedCode": lsd.get("derivedCode") or lsd.get("code") or "",
        "statusByLocale": lsd.get("statusByLocale") or "",
        "description": lsd.get("description") or "",
        "HasException": 1,
        "LatestEventTimestampUtc": "2025-10-18T12:07:41Z",
        # provide ancillary details to indicators
        "latestStatusDetail": lsd,
        # optional fallback if your indicators look at raw:
        "raw": SAMPLE_BODY,
        # ensure debug print in WorkbookProcessor has this optional column
        "ScanEventsCount": 0,
    }


@pytest.fixture
def patched_enricher(monkeypatch):
    # Patch Enricher.enrich to return the normalized row (with latestStatusDetail)
    from order_shipping_status.pipelines import enricher as enricher_mod

    def fake_enrich(self, df, *, sidecar_dir=None):
        return pd.DataFrame([_normalized_row_from_sample()])
    monkeypatch.setattr(enricher_mod.Enricher, "enrich", fake_enrich)
    return True


def test_damaged_detected_and_sheet_written(tmp_path: Path, patched_enricher):
    src = tmp_path / "input.xlsx"
    dst = tmp_path / "input.processed.xlsx"
    pd.DataFrame([{"Tracking Number": "seed"}]).to_excel(src, index=False)

    lg = logging.getLogger("test")
    lg.addHandler(logging.NullHandler())
    lg.propagate = False

    proc = WorkbookProcessor(logger=lg, enable_date_filter=False)
    proc.process(src, dst, env_cfg=None)

    wb = load_workbook(dst)
    assert "Damaged or Returned" in wb.sheetnames

    ws = wb["Damaged or Returned"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows, "Sheet should have headers"
    headers = list(rows[0])

    # accept several possible TN header variants
    tn_headers = [h for h in ("Tracking Number", "trackingNumber",
                              "tracking_number", "TrackingNumber") if h in headers]
    assert tn_headers, f"Tracking number column not found in headers: {headers}"
    tn_col = tn_headers[0]
    data = [dict(zip(headers, r)) for r in rows[1:]]
    matches = [r for r in data if str(r.get(tn_col)) == "394273834920"]
    assert matches, "Damaged TN not found on 'Damaged or Returned' sheet"

    # If Damaged column is present, expect it to be 1
    if "Damaged" in headers:
        assert all(int((r.get("Damaged") or 0)) == 1 for r in matches)
