from __future__ import annotations

from pathlib import Path
import logging
import pandas as pd
from openpyxl import load_workbook


# ---------- Test fixtures & helpers ----------

class _QuietLogger:
    def debug(self, *a, **k): pass
    def info(self, *a, **k): pass
    def warning(self, *a, **k): pass
    def error(self, *a, **k): pass


def _make_row_with(
    *,
    tn: str,
    code: str = "SE",
    status: str = "Shipment exception",
    desc: str = "Shipment exception",
    lsd: dict | None = None,
    raw: dict | None = None,
) -> dict:
    """
    Build a single normalized row your processor expects post-enrichment.
    Includes columns some debug prints rely on to avoid KeyError.
    """
    return {
        "Tracking Number": tn,
        "derivedCode": code,
        "statusByLocale": status,
        "description": desc,
        "HasException": 1,
        "LatestEventTimestampUtc": "2025-10-18T12:00:00Z",
        "ScanEventsCount": 0,
        "DaysSinceLatestEvent": 0,
        "latestStatusDetail": lsd or {},
        "raw": raw or {},
    }


def _read_sheet(path: Path, sheet: str) -> list[dict]:
    wb = load_workbook(path)
    assert sheet in wb.sheetnames, f"Expected '{sheet}' sheet"
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    assert rows, f"'{sheet}' should have headers"
    headers = list(rows[0])
    return [dict(zip(headers, r)) for r in rows[1:]]


# ---------- Tests ----------

def test_unable_to_deliver_detected_from_ancillary(tmp_path, monkeypatch):
    """
    If status is a shipment exception and ancillaryDetails include 'Unable to deliver',
    mark UnableToDeliver == 1 and ensure the row appears in All Issues.
    """
    # Build a minimal FedEx-shaped raw body with ancillary 'Unable to deliver'
    SAMPLE_BODY = {
        "output": {
            "completeTrackResults": [{
                "trackingNumber": "999111222333",
                "trackResults": [{
                    "latestStatusDetail": {
                        "code": "SE",
                        "derivedCode": "SE",
                        "statusByLocale": "Shipment exception",
                        "description": "Shipment exception",
                        "ancillaryDetails": [
                            {
                                "reasonDescription": "Unable to deliver - bad address",
                                "actionDescription": "Unable to deliver"
                            }
                        ],
                    },
                    "scanEvents": []
                }]
            }]
        }
    }
    lsd = SAMPLE_BODY["output"]["completeTrackResults"][0]["trackResults"][0]["latestStatusDetail"]

    # Patch Enricher.enrich to inject our normalized row
    from order_shipping_status.pipelines import enricher as enricher_mod

    def fake_enrich(self, df, *, sidecar_dir=None):
        row = _make_row_with(
            tn="999111222333",
            code="SE",
            status="Shipment exception",
            desc="Shipment exception",
            lsd=lsd,
            raw=SAMPLE_BODY,
        )
        return pd.DataFrame([row])
    monkeypatch.setattr(enricher_mod.Enricher, "enrich", fake_enrich)

    # Prepare input file (content is replaced by fake_enrich anyway)
    src = tmp_path / "in.xlsx"
    dst = tmp_path / "out.xlsx"
    pd.DataFrame([{"Tracking Number": "seed"}]).to_excel(src, index=False)

    # Run processor
    from order_shipping_status.pipelines.workbook_processor import WorkbookProcessor
    lg = logging.getLogger("test")
    lg.addHandler(logging.NullHandler())
    lg.propagate = False

    proc = WorkbookProcessor(logger=lg, enable_date_filter=False)
    proc.process(src, dst, env_cfg=None)

    # Validate output
    rows = _read_sheet(dst, "All Issues")
    row = next((r for r in rows if str(
        r.get("Tracking Number")) == "999111222333"), None)
    assert row, "TN row not found in All Issues"
    assert int(row.get("HasException", 0)) == 1
    # New flag
    assert int(row.get("UnableToDeliver", 0)) == 1
    # Damaged should not be set here
    assert int(row.get("Damaged", 0)) == 0


def test_unable_to_deliver_detected_from_status_or_description(tmp_path, monkeypatch):
    """
    If latestStatusDetail.ancillaryDetails are missing, but the text fields contain
    'Unable to deliver', we still mark UnableToDeliver == 1.
    """
    SAMPLE_BODY = {
        "output": {
            "completeTrackResults": [{
                "trackingNumber": "888777666555",
                "trackResults": [{
                    "latestStatusDetail": {
                        "code": "SE",
                        "derivedCode": "SE",
                        "statusByLocale": "Shipment exception",
                        "description": "Unable to deliver – recipient not available",
                        # No ancillaryDetails on purpose
                    },
                    "scanEvents": []
                }]
            }]
        }
    }
    lsd = SAMPLE_BODY["output"]["completeTrackResults"][0]["trackResults"][0]["latestStatusDetail"]

    from order_shipping_status.pipelines import enricher as enricher_mod

    def fake_enrich(self, df, *, sidecar_dir=None):
        row = _make_row_with(
            tn="888777666555",
            code="SE",
            status="Shipment exception",
            desc="Unable to deliver – recipient not available",
            lsd=lsd,
            raw=SAMPLE_BODY,
        )
        return pd.DataFrame([row])
    monkeypatch.setattr(enricher_mod.Enricher, "enrich", fake_enrich)

    src = tmp_path / "in2.xlsx"
    dst = tmp_path / "out2.xlsx"
    pd.DataFrame([{"Tracking Number": "seed"}]).to_excel(src, index=False)

    from order_shipping_status.pipelines.workbook_processor import WorkbookProcessor
    lg = logging.getLogger("test")
    lg.addHandler(logging.NullHandler())
    lg.propagate = False

    proc = WorkbookProcessor(logger=lg, enable_date_filter=False)
    proc.process(src, dst, env_cfg=None)

    rows = _read_sheet(dst, "All Issues")
    row = next((r for r in rows if str(
        r.get("Tracking Number")) == "888777666555"), None)
    assert row, "TN row not found in All Issues"
    assert int(row.get("HasException", 0)) == 1
    assert int(row.get("UnableToDeliver", 0)) == 1
    assert int(row.get("Damaged", 0)) == 0
